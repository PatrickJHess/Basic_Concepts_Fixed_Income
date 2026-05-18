# -*- coding: utf-8 -*-
"""
Module for Basic Concepts Of Fixed Income Volume
includes shared functions
    create_workbook
    file_graph
    fill_function
    one_y_axis
    parent_folder_subolder
    single_newton_raphson

"""

list_install=['numpy','pandas','openpyxl','pathvalidate',
             'matplotlib','python-dateutil','holidays',
            'ipynbname','bs4','pandas_market_calendars',
             'scipy','pandas-datareader','pandas<3.0']
content='\n'.join(list_install)
with open('requirements.txt','w') as file:
 file.write(content)
import subprocess
subprocess.run(['pip install -q -r requirements.txt'],shell=True,  stdout=subprocess.DEVNULL)

import requests
import pandas as pd
import numpy as np
import pandas_market_calendars as mcal
from pandas.tseries.holiday import GoodFriday, USFederalHolidayCalendar
from datetime import date,datetime
import calendar
from io import StringIO
from dateutil.relativedelta import relativedelta
import os
import sys
import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from IPython.display import display, Markdown as md, HTML
from scipy.optimize import minimize
import getpass
import pandas_datareader.data as web
import json
import time

def create_workbook(df,sheet_name='sheet1', save_config=None):
    """
    Writes a DataFrame to a specific sheet in an Excel workbook and auto-fits
    column widths for readability.

    Args:
        sheet_name (str): The name of the sheet to create or replace.
        df (pd.DataFrame): The DataFrame to write.
        save_config (dict, optional): Configuration for saving the file, passed
         Keys: 'volume':folder, 'chapter':'subfolder, 'file_name':file name. Defaults to {}.     
    """
    import os
    import re
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    from IPython.display import display, Markdown as md

    
    # Fix mutable default argument
    if save_config is None:
        save_config = {}

    # --- 1. Sanitize Sheet Name ---
    sane_sheet_name = re.sub(r'[\\*?:/\[\]]', '', str(sheet_name))
    if len(sane_sheet_name) > 31:
        sane_sheet_name = sane_sheet_name[:31]
    if not sane_sheet_name:
        sane_sheet_name = "Sheet1"

    # --- 2. Get Save Path ---
    file_name = save_config.get('file_name', 'output.xlsx')
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'
    save_config['file_name']=file_name
    # Assuming save_results is defined elsewhere in your code
    try:
        path_filename = save_results(save_config=save_config)
        if path_filename is None:
            path_filename = file_name
    except NameError:
        # Fallback if save_results is not defined
        path_filename = file_name

    # --- 3. Write and Format ---
    try:
        # Create empty workbook if it doesn't exist to allow 'append' mode
        if not os.path.exists(path_filename):
            pd.DataFrame().to_excel(path_filename, sheet_name="Sheet1")

        # Write DataFrame
        with pd.ExcelWriter(
            path_filename,
            mode='a',
            engine='openpyxl',
            if_sheet_exists='replace',
            datetime_format='YYYY-MM-DD'
        ) as writer:
            # FIXED: Using sane_sheet_name instead of sheet_name
            df.to_excel(writer, sheet_name=sane_sheet_name, index=True)

        # Format with openpyxl
        workbook = openpyxl.load_workbook(path_filename)
        
        # FIXED: Using sane_sheet_name
        try:
            ws = workbook[sane_sheet_name]
        except KeyError:
            print(f"Error: Sheet '{sane_sheet_name}' not found after writing.")
            return

        # Auto-fit columns
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in column_cells:
                try:
                    if cell.value: # Check if cell is not empty
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        # Delete default Sheet1 if necessary
        if 'Sheet1' in workbook.sheetnames and sane_sheet_name != 'Sheet1' and len(workbook.sheetnames) > 1:
            del workbook['Sheet1']
            
        workbook.save(path_filename)
        display(md(f"### ***✅ Successfully wrote and formatted sheet {sane_sheet_name} in {path_filename}***"))
        
    except Exception as e:
        display(md("### ❌ **ERROR during Excel write/format:**"))
        print(f"Exception details: {e}")
 
def save_results(save_config: dict = None):
    """
    Interactively prompts the user to confirm and generate a safe save path.

    This function pauses execution and asks the user (y/n) if they want to save
    a file. It is environment-aware:

    - **In Google Colab:** It attempts to use '/content/drive/MyDrive'. If not
      mounted, it will try to mount it. If mounting fails, it falls back
      to the temporary '/content' directory and issues a warning.
    - **In a local environment:** It uses the current working directory.

    The function constructs a full path from the base folder and the optional
    'volume' and 'chapter' subdirectories. All path components are sanitized
    using `pathvalidate`.

    Args:
        save_config (dict, optional): A dictionary containing path components.
            'volume' (str, optional): The name of a top-level subdirectory.
            'chapter' (str, optional): The name of a nested subdirectory.
            'file_name' (str, optional): The final file name.
                Defaults to 'output.txt' if not provided or if sanitization
                results in an empty string.

    Returns:
        str or None: A complete, sanitized, absolute string path to the
            file if the user confirms 'y' and path creation succeeds.
            Returns None if the user chooses 'n' or if an error occurs.
    
    Example:
        config = {
            'volume': 'My_Notebooks',
            'chapter': 'Chapter_01',
            'file_name': 'results.json'
        }
        save_path = save_results(config)
        
        # User inputs 'y'
        
        if save_path:
            # save_path might be '/content/drive/MyDrive/My_Notebooks/Chapter_01/results.json'
            print(f"Saving to: {save_path}")
            # ... proceed to write file ...
    """
    import os
    import sys
    from pathlib import Path
    from pathvalidate import sanitize_filepath
    from IPython.display import display, Markdown as md

    save_config = save_config or {}

    # --- Handle IPython imports safely ---
    is_colab = 'google.colab' in sys.modules
    try:
        from IPython.display import display, Markdown as md
        def display_msg(text): display(md(text))
    except ImportError:
        def display_msg(text): print(text.replace('###', '').replace('*', ''))

    # --- 1. Get user's choice (y/n) ---
    default_value = "n"
    prompt = "❓ Do you want to save the file? (y/n) (press enter for n): "

    try:
        raw_input = input(prompt).strip().lower()
        choice = raw_input if raw_input else default_value
    except (EOFError, KeyboardInterrupt):
        return None

    while choice not in ['y', 'n']:
        choice = input(prompt).strip().lower()

    # --- 2. Handle "No" ---
    if choice == 'n':
        display_msg('### ***❌ File Not Saved***.')
        return None

    # --- 3. Handle "Yes" ---
    display_msg("### ***⌛ Generating A Path***")

    # --- Get and sanitize config values ---
    volume = sanitize_filepath(save_config.get('volume', ''))
    chapter = sanitize_filepath(save_config.get('chapter', ''))
    
    # The caller should provide the extension, but we fall back to a generic name just in case
    file_name = sanitize_filepath(save_config.get('file_name', 'output'))
    if not file_name:
        file_name = 'output'

    subfolder = os.path.join(volume, chapter)

    # --- Drive/Folder Logic ---
    base_folder = ''
    drive_path = '/content/drive/MyDrive'
    
    if is_colab:
        if os.path.exists(drive_path):
            base_folder = drive_path
        else:
            try:
                from google.colab import drive
                drive.mount('/content/drive')
                base_folder = drive_path
            except Exception as e:
                base_folder = '/content'
                display_msg(f"### ⚠️ **Drive Mount Failed:** {e}. Saving to temporary '/content' folder.")
    else:
        base_folder = os.getcwd()

    # --- Path Creation ---
    try:
        full_folder_path = os.path.join(base_folder, subfolder)
        path_obj = Path(full_folder_path)
        path_obj.mkdir(parents=True, exist_ok=True)
        
        final_path_str = os.path.join(str(path_obj), file_name)

        display_msg(f'### ✅ **File Path Generated:**\n`{final_path_str}`')
        
        if base_folder == '/content':
            display_msg('### ⚠️ *File is in a temporary location and will be lost on runtime restart.*')

        return final_path_str

    except Exception as e:
        display_msg('### ❌ **ERROR Creating Directory:**')
        print(e)
        return None

def file_graph(save_config, title):
  """
  Generates a sanitized file path and name for saving a graph.

  This function intelligently determines the correct base path (handling Google
  Colab vs. local environments), constructs a full path from a configuration
  dictionary, creates the directory if it doesn't exist, and returns a
  sanitized path and filename.

  Args:
    save_config (dict): A dictionary containing save path components like
                        'Parent', 'Folder', 'Subfolder', and 'File'.
    title (str): The title of the graph, used as a fallback for the filename.

  Returns:
    tuple: A tuple containing the (path, file_name) if successful.
    None: If the save_config dictionary is empty.
  """
  import os
  import sys
#Empty save_config dictionary
  if not save_config:
    print(f'⚠️ No Config info...Graph Not Saved')
    return None
  from pathvalidate import sanitize_filepath,sanitize_filename
# Set a default title if none is provided
  if not title:
    title='Add Title'

  # Determine the execution environment (Colab or Local) 
  # Check if the code is running in a Google Colab notebook.
  is_colab='google.colab' in sys.modules
  if is_colab:
    if not os.path.exists('/content/drive/My Drive'):
      from google.colab import drive
      drive.mount('/content/drive')
    default_parent='/content/drive/MyDrive'
  # On a local machine, the default save location is the current working directory.
  else:
      default_parent = os.getcwd()

  # Construct the directory path from config
  # Get and sanitize path components using .get() for safety
  parent = sanitize_filepath(save_config.get('Parent', default_parent))
  folder = sanitize_filepath(save_config.get('Folder', ''))
  # 2. FIX: Corrected case for 'SubFolder'
  subfolder = sanitize_filepath(save_config.get('Subfolder', ''))

  # Join the components to form the full directory path.
  path = os.path.join(parent, folder, subfolder)

  # Create the directory if it doesn't exist
  # The 'exist_ok=True' flag prevents an error if the directory is already there.
  os.makedirs(path, exist_ok=True)
 
  #Determine the filename
  # Check if a specific filename is provided in save_config
  if not save_config.get('File',''):
    # If not, create a filename from the graph title with a .pdf extension
    file_name=title+'.pdf'
  else:
    # Otherwise, use the filename from the config.
    file_name=save_config['File']
  # Sanitize the final filename to remove any invalid characters.
  file_name=sanitize_filename(file_name)
  #Return the sanitized path and filename
  return path,file_name


def fill_function_refactored(fill_dict, x):
    """
    Parses a dictionary to get configuration for filling a plot area.

    Args:
        fill_dict (dict): Dictionary with keys like 'Color', 'Alpha', 'Start', 'End'.
        x (list or array): The data array used to determine valid boundaries.

    Returns:
        tuple: A tuple containing (start, end, alpha, color, label).
    
    Raises:
        ValueError: If the provided 'Start' or 'End' values are out of bounds.
    """
    # Use the .get() method for clean defaults
    color = fill_dict.get('Color', 'green')
    alpha = fill_dict.get('Alpha', 0.25)
    label = fill_dict.get('Label', 'Not Defined')
    
    # Get start and end values, defaulting to the full range
    start = fill_dict.get('Start', 0)
    end = fill_dict.get('End', len(x))

    # Validate the start and end values *after* they are defined
    if not (0 <= start < len(x)):
        raise ValueError(f"Start value '{start}' is out of bounds for data of length {len(x)}.")
    
    if not (0 < end <= len(x)):
        raise ValueError(f"End value '{end}' is out of bounds for data of length {len(x)}.")

    return start, end, alpha, color, label



def one_y_axis(x_data, y_data_list, title, series_labels, xlabel, ylabel,
                       markers, figure_size, y_limits,save_config={}, fill_config={},
                       colors=None):
    '''
    Plots data on a single y-axis.

    Args:
        x_data (array-like): Data for the x-axis.
        y_data_list (list of array-like): A list of datasets for the y-axis.
        title (str): The title of the graph.
        series_labels (list of str): Identifiers for each data series in the legend.
        xlabel (str): The label for the x-axis.
        ylabel (str): The label for the y-axis.
        markers (list of str): The markers to use for each series.
        figure_size (tuple): The width and height of the figure in inches.
        y_limits (tuple): The minimum and maximum values for the y-axis.
        save_config (dict, optional): Configuration for saving the file, passed
            to save_results(). Keys: 'volume', 'chapter', 'file_name'. Defaults to {}.
        fill_config (dict, optional): Configuration for filling areas.
            Keys: 'Between' (list of 1 or 2 indices from y_data_list),
                  'Start' (int, start index), 'End' (int, end index),
                  'Colors' (str), 'Labels' (str), 'Alpha' (float).
            Defaults to {}.
    Raises:
        ValueError: If input lists for series, markers, or colors do not match the number of y-datasets.
    '''
    import numpy as np
    from matplotlib import pyplot as plt
    num_series = len(y_data_list)
    # --- Input Validation ---
    if not all(len(lst) == num_series for lst in [series_labels, markers]):
        raise ValueError("The 'series_labels' and 'markers' lists must have the same length as 'y_data_list'.")

    if colors and len(colors) != num_series:
        raise ValueError("The 'colors' list must have the same length as 'y_data_list'.")

    # --- Plotting Setup ---
    fig = plt.figure(figsize=figure_size)
    fig.suptitle(title)
    plt.style.use('ggplot')

    if colors is None:
        # Generate a default color cycle if none are provided
        colors = plt.cm.viridis_r(np.linspace(0, 1, num_series))

# --- Plot Data Series ---
    for i in range(num_series):
        plt.plot(x_data, y_data_list[i], label=series_labels[i], marker=markers[i], color=colors[i])

    # --- Handle Fill Area ---
    if fill_config.get('Between'):
        if len(fill_config['Between']) > 2:
            raise ValueError("The 'Between' key in fill_config can contain a maximum of two indices.")


        # Get values from fill_config dict, providing safe defaults
        start = fill_config.get('Start', 0)
        end = fill_config.get('End', len(x_data))
        alpha = fill_config.get('Alpha', 0.3)
        color = fill_config.get('Colors', 'gray')
        label = fill_config.get('Labels', None) # 'None' won't create a legend item

        if len(fill_config['Between']) == 2:
            y1_index, y2_index = fill_config['Between']
            plt.fill_between(x_data[start:end],
                             y_data_list[y1_index][start:end],
                             y_data_list[y2_index][start:end],
                             color=color, alpha=alpha, label=label)
        else:
            y_index = fill_config['Between'][0]
            # Fills between the series and y=0
            plt.fill_between(x_data[start:end],
                             y_data_list[y_index][start:end],
                             color=color, alpha=alpha, label=label)

    # --- Final Touches ---
    plt.ylim(y_limits)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.legend()
    plt.tight_layout()

    # --- Save Figure ---
    # Calls the save_results function (assumed to be defined)

    path = save_results(save_config=save_config)
    if path:
     plt.savefig(path, dpi=300, bbox_inches='tight')

    plt.show()

def validate_date(datetime_obj):
  """
  Validates input is a date/datetime and returns a date object.

  This helper function checks if the input is either a 
  datetime.datetime or datetime.date object. If it's a 
  datetime.datetime, it converts it to a datetime.date.
  It always returns a datetime.date object.

  Args:
    datetime_obj (datetime.date or datetime.datetime): 
      The input to validate and convert.

  Returns:
    datetime.date: The resulting date object.

  Raises:
    TypeError: If the input is not a datetime.date or 
               datetime.datetime object.
  """
  from datetime import datetime, date
  
  #Validation ---
  
  # Check if the object is of the expected type (date or datetime)
  if not isinstance(datetime_obj, (datetime, date)):
      raise TypeError("Input must be a datetime or date object.")
  
  #Standardization ---
  
  # If the object is a datetime, convert it to a date.
  # This makes the return type consistent.
  if isinstance(datetime_obj, datetime):
    datetime_obj = datetime_obj.date()
  
  # Return the object, which is now guaranteed to be a date object.
  return datetime_obj


def last_day_month(date_value):
  """
  Calculates the last day of the month for a given date or datetime object.

  This function accepts either a datetime.datetime or datetime.date
  object and always returns a datetime.date object representing the
  final day (e.g., 30, 31, 28, or 29) of that month.

  Args:
    date_value (datetime.date or datetime.datetime): 
      The input date or datetime to find the month-end for.

  Returns:
    datetime.date: A date object for the last day of the given month.

  Raises:
    calls validate_date to check correct data
    TypeError: If the input is not a datetime.date or datetime.datetime.
  """
  from datetime import datetime,date
  import calendar
 
  #Validate the data- datetime_obj and settlement
  #If necessary convert datetime to date
  date_value=validate_date(date_value)
 
  # calendar.monthrange() returns a tuple: (weekday_of_first_day, num_days_in_month)
  # We only need the second value [1], which is the last day of the month.
  # Example: For Feb 2024 (leap year), it returns (3, 29)
  # Example: For Feb 2025, it returns (5, 28)
  month_end_day=calendar.monthrange(date_value.year,date_value.month)[1]

  return date(date_value.year,date_value.month,month_end_day)    

def settle_year_mat(maturity, settlement=None):
  """
  Returns a new date with the maturity month/day in the settlement year,
  and a boolean indicating if the original maturity was a month-end.

  Args:
    maturity (datetime.date or datetime.datetime): The original maturity date.
    settlement (datetime.date or datetime.datetime, optional):
      The settlement date. Defaults to today's date if None.

  Returns:
     tuple: A (datetime.date, bool) tuple containing:
       - The new maturity date (a date object).
       - True if the original maturity was the last day of its month, False otherwise.
  """
  from datetime import datetime,date
  import calendar

  #Validate the data- maturity_date and settlement
  #If necessary convert datetime to date
  
  #maturity
  maturity=validate_date(maturity)  

  #settlement
  if settlement is None:
      settlement = date.today()
  else:
    settlement=validate_date(settlement)
 
  # Determine year for settlement
  year=settlement.year

  # Get the total number of days in the given month and year
  days_in_month = calendar.monthrange(maturity.year, maturity.month)[1]

  # Check if the date's day is the last day of the month
  end_month = (maturity.day == days_in_month)

  if end_month:day=calendar.monthrange(year,maturity.month)[1]
  else:day=maturity.day
  return date(year,maturity.month,day),end_month

def next_last_coupon_dates(maturity, settlement=None, freq=2):
  """
  Calculates the next and last coupon dates relative to a settlement date.

  Args:
      maturity (datetime): The maturity date of the instrument.
      settlement (datetime, optional): The date for which the calculation is made.
                                      Defaults to the current date if None.
      freq (int, optional): The frequency of payments per year.
                            Valid values: 1 (annual), 2 (semi-annual),
                            4 (quarterly), 12 (monthly). Defaults to 2.

  Returns:
      tuple: A tuple containing (next_coupon_date, last_coupon_date).
  """

  import calendar
  from datetime import datetime, date
  from dateutil.relativedelta import relativedelta

  #Validate data
  #freq
  if freq not in [1, 2, 4, 12]:
      freq = 2 # Default to semi-annual if input is invalid
  
  #maturity
  maturity=validate_date(maturity)  

  #settlement
  if settlement is None:
      settlement = date.today()
  else:
    settlement=validate_date(settlement)

  num_months = 12 // freq

  #Determine Coupon Anniversary and Month-End Status ---
  # Start by finding the coupon date in the settlement year
  coupon_date, is_month_end = settle_year_mat(maturity, settlement=settlement)
 
  #Find the NEXT Coupon Date ---
  # If the anniversary in the settlement year has already passed, find the first one after
  if coupon_date < settlement:
      while coupon_date < settlement:
          coupon_date += relativedelta(months=num_months)
          if is_month_end:
              coupon_date = last_day_month(coupon_date)
  # If the anniversary is too far in the future, step back
  else:
      while (coupon_date - relativedelta(months=num_months)) >= settlement:
            coupon_date -= relativedelta(months=num_months)
            if is_month_end:
              coupon_date = last_day_month(coupon_date)

  next_coupon = coupon_date

  #Find the LAST Coupon Date ---
  # The last coupon is simply one period before the next coupon
  last_coupon = next_coupon - relativedelta(months=num_months)
  if is_month_end:
      last_coupon = last_day_month(last_coupon)

  return next_coupon, last_coupon


def convert_isda(settlement, prev_coupon):


    # make sure settlement is date
    # if not pandas timestamp convert and make it a date
    def make_date(date_value):
      if not isinstance(date_value,(datetime,date)):
        date_value=pd.Timestamp(date_value).date()
        return date_value
      # convert timestamps and datetimes to date
      else:
        try:
            return date_value.date()
        except:        
           return date_value

    settlement=make_date(settlement)
    prev_coupon=make_date(prev_coupon)


    if prev_coupon.year == settlement.year:
        days_in_year = 366 if calendar.isleap(settlement.year) else 365
        actual_days = (settlement - prev_coupon).days
        return actual_days / days_in_year

    # ISDA Split: Pivot on Jan 1st of the second year
    pivot_date = date(settlement.year, 1, 1)

    # All days from the coupon up to midnight Jan 1st
    days_prev = (pivot_date - prev_coupon).days
    days_prev_year = 366 if calendar.isleap(prev_coupon.year) else 365
    prev_ratio = days_prev / days_prev_year

    # All days from Jan 1st to the accrual end
    days_settlement = (settlement- pivot_date).days
    days_in_settlement_year = 366 if calendar.isleap(settlement.year) else 365
    settlement_ratio = days_settlement / days_in_settlement_year

    return prev_ratio + settlement_ratio

def _30_360_(settlement,prev_coupon):


    # make sure settlement is date
    # if not pandas timestamp convert and make it a date
    def make_date(date_value):
      if not isinstance(date_value,(datetime,date)):
        date_value=pd.Timestamp(date_value).date()
        return date_value
      # convert timestamps and datetimes to date
      else:
        try:
            return date_value.date()
        except:        
           return date_value


    settlement=make_date(settlement)
    prev_coupon=make_date(prev_coupon)

    # initalize
    number_days=0

    # accounting for years
    number_days+=(settlement.year-prev_coupon.year)*360

    # accounting for months
    number_days+=(settlement.month-prev_coupon.month)*30

    # accounting for days
    # February is the exception
    # Function to check if a date is the last day of its month (Feb 28 or 29)
    def feb_end(date_value):
        # If month is 2, check if it's the last day using calendar.monthrange
        return date_value.month == 2 and date_value.day == calendar.monthrange(date_value.year,
                                                                               date_value.month)[1]
    # STEP 1: Normalize Previous Coupon Date
    # Rule: If it's the 31st OR the end of Feb, it's 30.
    if prev_coupon.day == 31 or feb_end(prev_coupon):
        prev_coupon_days = 30
    else:
        prev_coupon_days = prev_coupon.day

    # STEP 2: Conditional Settlement Date
    # Rule: If it's the 31st OR the end of Feb AND the start was 30, it's 30.
    if (settlement.day == 31 or feb_end(settlement)) and prev_coupon_days == 30:
        settlement_days = 30
    else:
        settlement_days = settlement.day

    number_days+=settlement_days-prev_coupon_days

    return number_days/360

def accrued_interest(maturity, coupon, day_type='Actual/Actual', settlement=None, freq=2):
    """
      Returns the accrued interest for a bond.  Returns the accrued interest for a bond.

       Returns the accrued interest for a bond.  Returns the accrued interest for a bond.

    Args:
        maturity (datetime.date,timestamp, or datetime64): The maturity date of the bond.
        coupon (float): The annual coupon rate (e.g., 0.05 for 5%).
        day_types:
            Actual/Actual, Actual/365, Actual/360. and 30/360.
        settlement ((datetime.date,timestamp, or datetime64), optional):
         The settlement date. Defaults to today.
        freq (int, optional):
        Coupon frequency per year: 1 (annual). 2 (semi-annual)
                                    4 (quarterly), 12 (monthly)
                                    Defaults to 2.
    """
    from datetime import datetime,date
    import calendar
    from dateutil.relativedelta import relativedelta
    #Validate Data
    def validate_date(datetime_object):
      # check for datetime or date
      if not isinstance(datetime_object, (datetime, date)):
          raise TypeError("Input must be a datetime or date object.")
      # convert datetime to date
      if isinstance(datetime_object, datetime):
        datetime_object = datetime_object.date()
      return datetime_object
    maturity = validate_date(maturity)

    if settlement is None:
        settlement = date.today()
    else:
        settlement = validate_date(settlement)

    if freq not in [1, 2, 4, 12]:
        print(f"⚠️ Warning: Freq {freq} invalid. Assumed Semi-Annual (2).")
        freq = 2

    try:
        coupon = float(coupon)
        if coupon < 0: raise ValueError
    except:
        raise ValueError("Coupon must be a positive number.")

    # Define strategic_date
    # Get all the bond's potential payment dates in the next year
    mat_is_last=maturity.day==calendar.monthrange(maturity.year,maturity.month)[1]
    if mat_is_last:
      lastDay=calendar.monthrange(settlement.year+1,maturity.month)[1]
      strategic_date=date(settlement.year+1,maturity.month,lastDay)
    else:
     strategic_date=date(settlement.year+1,maturity.month,maturity.day)

    #The strategic date: minimum of actual and next year's maturity date
    strategic_date=min(maturity,strategic_date)
    pay_dates = scheduled_pay_dates(strategic_date, settlement=settlement, freq=freq)

    # Should sorted but check
    pay_dates.sort()

    # The first date after the settlement date is the next coupon date
    next_coupon = None
    for d in pay_dates:
        if d >= settlement:
            next_coupon = d
            break

    # Bond has matured or annual coupon is zero
    if next_coupon is None or coupon==0:
        return 0.0

    #Calculate Previous Coupon Date
    num_months = int(12 // freq)
    prev_coupon = next_coupon - relativedelta(months=num_months)

    # Check for Month End adjustment on the calculated previous date
    is_next_month_end = next_coupon.day == calendar.monthrange(maturity.year, maturity.month)[1]

    if is_next_month_end:
        last_day_of_prev_month = calendar.monthrange(prev_coupon.year, prev_coupon.month)[1]
        prev_coupon = date(prev_coupon.year, prev_coupon.month, last_day_of_prev_month)

    # The day a coupon is paid is also the first day of the new cycle.

    accrued_value = 0.0

    if day_type == 'Actual/Actual':
        days_since_last = (settlement - prev_coupon).days
        days_between = (next_coupon - prev_coupon).days
        #  (DaysHeld / DaysInPeriod)
        accural_ratio= days_since_last/days_between
        # Actual/Actual uses the coupon paid on the date
        accrued_value = (coupon / freq) * accural_ratio


    elif day_type == '30/360':
        accural_ratio =_30_360_(settlement,prev_coupon)
        # Formula: Coupon * accrural ratio
        accrued_value = coupon * accural_ratio

    elif day_type == 'Actual/360':
        days_since_last = (settlement - prev_coupon).days
        accural_ratio= days_since_last/360
      # Formula: Coupon * accrural ratio        
        accrued_value = coupon * accural_ratio

    elif day_type == 'Actual/365':
        accural_ratio = convert_isda(settlement,prev_coupon)
      # Formula: Coupon * accrural ratio        
        accrued_value = coupon * accural_ratio

    else:
        # Fallback
        print(f"⚠️ Warning: Unknown day_type {day_type}. Using Actual/Actual.")
        days_since_last = (settlement - prev_coupon).days
        days_between = (next_coupon - prev_coupon).days
        #  (DaysHeld / DaysInPeriod)
        accural_ratio= days_since_last/days_between
        # Actual/Actual uses the coupon paid on the date
        accrued_value = (coupon / freq) * (days_since_last / days_between)

    return accrued_value



def scheduled_pay_dates(last_date,settlement=None,freq=2):
  '''
    Generates a chronological list of coupon payment dates from settlement to last_date.
    The function calculates dates backward from the last_date date based on the
    specified frequency. It handles standard bond market "end-of-month" logic:
    if the last_date date is the last day of a month, all preceding coupon payments
    are snapped to the last day of their respective months.
    Args:
        last_date (datetime.date,timestamp, or datetime64): The last date checked.
            Accepts a date object.
        settlement (datetime.date,timestamp, or datetime64): The settlement date (start of analysis).
            Coupons falling before this date are excluded. Defaults to date.today().
        freq (int, optional): The number of coupon payments per year.
            Accepted values:
            * 1: Annual, 2: Semi-Annual (Default), 4: Quarterly, 12: Monthly

    Returns:
        list[datetime.date]: A list of coupon dates sorted chronologically
        (earliest to latest), ending with the last_date date..
  '''



# convert settlement and last_date to datetime.date
  def make_date(date_value):
    if not isinstance(date_value,(datetime,date)):
      date_value=pd.Timestamp(date_value).date()
      return date_value
    # convert timestamps and datetimes to date
    else:
      try:
          return date_value.date()
      except:        
         return date_value


  #Validate the data- last_date, coupon, settlement, freq
  #last_date
  last_date=make_date(last_date)

  #settlement
  if settlement is None:
      settlement = date.today()
  else:
      settlement=make_date(settlement)
  #freq
  if int(freq) not in [1,2,4,12]:
      display(md(f"### ⚠️  your assigned freq {freq} it must be (1, 2, 4, or 12)\
     \n     semi-annual assumed (2)."))
      freq=int(2)

  # check last_date greater than settlement
  if last_date<=settlement:
    raise ValueError("last_date must be greater than the settlement date")

  # calculate the number of months between each coupon payment.
  num_months=int(12/freq)
 
  # timestamps so hat we can use pandas offset
  last_day=pd.Timestamp(last_date)
  settlement=pd.Timestamp(settlement)
  # raw_dates is a pandas datetime index that starts and goes 60 times
  # need to check for month_end and respect it (MonthsEnd)
  if last_day.is_month_end:
    raw_dates = pd.DatetimeIndex([last_day - pd.offsets.MonthEnd(num_months * i) for i in range(360)])
  else:
    raw_dates = pd.DatetimeIndex([last_day - pd.DateOffset(months=num_months * i) for i in range(360)])

  # only include dates greater than or equal to settlement date
  valid_dates=raw_dates[raw_dates>=settlement]

  # convert timestamps back to date and make the order chronological with splice
  valid_dates=pd.to_datetime(valid_dates).date

  return valid_dates[::-1]




def apply_market_pay_dates(df, date_col='Date', calendar_name='SIFMAUS'):
  """
    Maps a set of input dates to the next valid market settlement/trading day.

    This function takes a scalar date, a list of dates, or a DataFrame containing 
    a date column, and determines the nearest open market day on or after each 
    date using a specified pandas_market_calendars exchange calendar.

    Args:
        df (pd.DataFrame, list-like, or scalar): The input data containing dates.
        date_col (str, optional): The name of the date column. Defaults to 'Date'.
        calendar_name (str, optional): The market calendar. Defaults to 'SIFMAUS'.

    Returns:
        pd.DataFrame: A DataFrame sorted by the date column, containing the 
            original dates and a new 'Settlement' column with adjusted market dates.
  """
  # imports required in custom module
  import pandas as pd
  from datetime import date, datetime
  import pandas_market_calendars as mcal

  # if not a Dataframe convert
  if isinstance(df, pd.DataFrame):
    if date_col not in df.columns:
      raise KeyError(f"Column '{date_col}' not found in DataFrame.")
  elif pd.api.types.is_list_like(df) and not isinstance(df, (str, bytes)):
    df = pd.DataFrame(df, columns=[date_col])
  elif pd.api.types.is_scalar(df):
    df = pd.DataFrame([df], columns=[date_col])
  else:
    raise TypeError(f"Expected a DataFrame, scalar, or list-like object, got: {type(df).__name__}")

  # ensure date_col is datetime and sorted (required for merge_asof)
  try:
    df[date_col] = pd.to_datetime(df[date_col])
  except Exception as e:
    raise TypeError(f"Could not convert values in '{date_col}' to datetime. Details: {e}")
         
  df = df.sort_values(date_col)

  # define dynamic range with a safety buffer (e.g., +14 days for the end)
  start_dt = df[date_col].min()
  end_dt = df[date_col].max() + pd.Timedelta(days=14)

  # fetch calendar and generate valid days
  cal = mcal.get_calendar(calendar_name)
  valid_days = cal.valid_days(start_date=start_dt, end_date=end_dt)
  # remove timeszone info and set to start of day
  valid_days = pd.DataFrame({'Settlement': pd.to_datetime(valid_days.tz_localize(None).normalize())})
  df[date_col] = df[date_col].astype('datetime64[ns]')
  valid_days['Settlement'] = valid_days['Settlement'].astype('datetime64[ns]')
  # merge the Dataframes
  # direction='forward' means: If today is closed, take the next open day.
  df = pd.merge_asof(
      df,
      valid_days,
      left_on=date_col,
      right_on='Settlement',
      direction='forward'
  )

  return df

def adjust_bond_pay_dates(dates,calendar='SIFMAUS'):
  """
  Adjusts bond payment dates to account for holidays and weekends.
  dates can be a scalar, pandas series, or numpy array (datetime.date,timestamp, or datetime64)
  """

  import pandas as pd
  import numpy as np
  import pandas_market_calendars as mcal
  from pandas.tseries.holiday import GoodFriday

  # Ensure dates is a DatetimeIndex
  if not pd.api.types.is_scalar(dates):
      dates = pd.DatetimeIndex(pd.to_datetime(dates))
  else:
      dates = pd.DatetimeIndex(pd.to_datetime([dates]))
      
  sifma = mcal.get_calendar(calendar)
  sifma_holidays = sifma.holidays().holidays
  good_fridays = GoodFriday.dates('2000-01-01', '2060-12-31')

  # Convert to DatetimeIndex and use .union() (which automatically deduplicates and sorts)
  sifma_idx = pd.DatetimeIndex(sifma_holidays)
  gf_idx = pd.DatetimeIndex(good_fridays)
  master_bond_holidays = sifma_idx.union(gf_idx)

  # Create a CustomBusinessDay offset using the combined holidays
  numpy_holidays = master_bond_holidays.values.astype('datetime64[D]')

  # Use NumPy for lightning-fast, fully vectorized date math (No warnings!)
  actual_payment_dates = np.busday_offset(
      dates.values.astype('datetime64[D]'),
      offsets=0,
      roll='forward',
      holidays=numpy_holidays
  )
  
  final_dates = pd.to_datetime(actual_payment_dates).date

  return final_dates
  
def bond_pay_data(maturity, coupon, settlement=None, freq=2):
    '''
    Function calculates payment Dates And Amounts.
    maturity is a datetime object and coupon is a real number.
    Required arguments are maturity and annual coupon.
    If provided, the value of settlement is a datetime object;
    otherwise defaults to date.today()
    freq defaults to semi-annual but accepts freq equal
    to 1 for annual, 2 for semi-annal, 4 for quarterly, and 12 for monthly.
    The function assumes a par value of 100.
    Returns Numpy arrays of dates and amounts.

    Raises:
        TypeError: If maturity or settlement are not datetime objects.
        ValueError: If inputs are not logically valid (e.g., negative coupon,
                    maturity before settlement).
    '''
    from datetime import datetime, date
    from dateutil.relativedelta import relativedelta
    import pandas as pd
    import numpy as np
    from IPython.display import display, Markdown as md

    # Validate the data - maturity, coupon, settlement, freq
    def make_date(date_value):
      # datetime64 are conerted
      if not isinstance(date_value,(datetime,date)):
        try:
          date_value=pd.Timestamp(date_value).date()
        except Exception as e: # Catch anything else unexpected
          print(f"wrong type for settlement or maturity {e}")
  
        date_value=pd.Timestamp(settlement).date()
      # convert timestamps and datetimes to date
      else:
        try:
          date_value=date_value.date()
        except:
          pass
      return date_value

    # maturity
    maturity = make_date(maturity)

    # settlement
    if settlement is None:
        settlement = date.today()
    else:
        settlement = make_date(settlement)

    # coupon
    try:
        coupon = float(coupon)
        if coupon < 0:
            raise ValueError("coupon rate cannot be negative.")
    except (ValueError, TypeError):
        raise ValueError("coupon must be a valid number.")

    # freq
    if int(freq) not in [1, 2, 4, 12]:
        display(md(f"### ⚠️ your assigned freq {freq} it must be (1, 2, 4, or 12)\n ### semi-annual assumed (2)."))
        freq = int(2)

    # check maturity greater than settlement
    if maturity <= settlement:
        raise ValueError("maturity must be greater than the settlement date")

    if coupon == 0:
        # Adjust maturity for non-settlement day and return date and face value
        adjust_maturity = adjust_bond_pay_dates(maturity)
        return np.array(adjust_maturity), np.array([100.0])

    # get scheduled payment dates from helper function scheduled_pay_dates
    scheduled_dates = scheduled_pay_dates(maturity, settlement, freq)

    # Pandas DataFrame Settlement desired column
    pay_dates = adjust_bond_pay_dates(scheduled_dates)

    # calculate payments
    # coupon divided by freq at each date
    pay = np.full(len(pay_dates), coupon / freq)

    # Add principal payment as last cash payment
    pay[-1] += 100

    return pay_dates,pay

def create_payoff_df(df, settlement,OLS=False):
    adjusted_maturities = adjust_bond_pay_dates(list(df.index))
    all_maturities = set(adjusted_maturities)

    df_payoff_columns = sorted(all_maturities)
    df_payoff_index=[i for i in range(len(df.index))]

    df_payoff = pd.DataFrame(
        np.zeros((len(df), len(df_payoff_columns))),
        columns=df_payoff_columns,
        index=df_payoff_index
    )
    total_rows = len(df)
    # Define a clean, pleasing HTML template for our status box
    def status_box(current, total):
        return f"""
        <div style="font-family: Arial, sans-serif; padding: 10px 15px; background-color: #f8f9fa; 
                    border-left: 4px solid #007bff; border-radius: 4px; width: fit-content; color: #333;">
            <b style="color: #007bff;">⚙️ Processing Bonds:</b> {current} of {total} added to DataFrame
        </div>
        """
    
    # initial display showing 0 bonds added
    progress_ui = display(HTML(status_box(0, total_rows)), display_id=True)

    for index,(maturity, coupon) in enumerate(zip(df.index, df['Coupon'])):

        # bond_pay_data returns payment dates and amounts
        row_pay_data = bond_pay_data(maturity, coupon, settlement=settlement)

        # Find any dates that aren't already columns
        new_dates = set(row_pay_data[0].flatten()) - all_maturities

        if new_dates:
          if OLS:
            df_clean = df_payoff.loc[(df_payoff != 0).any(axis=1),
                                         (df_payoff != 0).any(axis=0)]
            print("✅ DataFrame Complete (Exited Early)!")
            return df_clean
          else:
            # ✅ FIX: Add new dates to our master set and reindex
            all_maturities.update(new_dates)
            df_payoff = df_payoff.reindex(columns=sorted(all_maturities), fill_value=0.0)
 
        #    fill up the columns
        df_payoff.loc[index, row_pay_data[0]] = row_pay_data[1]
         # update the progress bar
        progress_ui.update(HTML(status_box(index + 1, total_rows)))
    # re-sort the columns so dates are chronological
    df_payoff = df_payoff.reindex(sorted(df_payoff.columns), axis=1)
    progress_ui.update(HTML("""
        <div style="font-family: Arial, sans-serif; padding: 10px 15px; background-color: #e6f4ea; 
                    border-left: 4px solid #34a853; border-radius: 4px; width: fit-content; color: #137333;">
            <b>✅ DataFrame Complete!</b> All bonds added successfully.
            </div>
            """))
    return df_payoff


def ns_spot_rates(interim_estimates,mat_years):

  # current values of estimates
  b_0,b_1,b_2,tau=interim_estimates

  # t saves typing
  t=mat_years

  # Avoid division by zero for t=0
  t = np.where(t == 0, 1e-8, t)
  # Nelson-Siegel model
  spot_rates=b_0+b_1*(1-np.exp(-t/tau))/(t/tau)\
  +b_2*((1-np.exp(-t/tau))/(t/tau)-np.exp(-t/tau))

  # pass these rates to the objective function for step one
  return spot_rates,t

def estimate_ns_parameters(df_payoff_matrix,P_actual,mat_years,guesses):

  # objective function is sum squared residuals (SSR)
  def predict_prices(interim_estimates,df_payoff_matrix,P_actual,mat_years):

    # for step one get rates
    spot_rates,mat_years=ns_spot_rates(interim_estimates,mat_years)

    # calculate zero prices
    zero_prices=np.exp(-spot_rates*mat_years)

    # calculate predicted prices (@ here is the same as np.matmul)
    P_predicted=df_payoff_matrix@zero_prices

    # step 2 calculate the distance with sum of squared residuals
    ssr= np.sum((P_actual.values-P_predicted)**2)
    return ssr

  # use the scipy minimize function to estimate parameters

  # Nelder-Mead doesn't take derivatives and tolerant of data
  method='Nelder-Mead'

  # minimize results
  ns_results=minimize(fun=predict_prices,
                  x0=guesses,
                  args=(df_payoff_matrix,P_actual,mat_years),
                  method=method)

  # get estimated coefficients of Nelson-Siegel
  b_0,b_1,b_2,tau=ns_results.x

  # get status of minimization
  completion_status=ns_results.message

  display(f'Completion status: {completion_status}')
  display(f'Long Rate (Beta Zero) {b_0:.4f}..Slope (Beta One) {b_1: .4f}...\
  Shape (Beta Three) {b_2: .4f}...Scaling (Tau) {tau: .4f}') 
  return ns_results

  

  
def calc_par_yield(maturity,df=None,freq=2,
                   bond_type='Government',settlement=None):
  """
  Calculates the par yield for a bond given its characteristics and market zero-coupon prices.

  Args:
    maturity (datetime): The maturity date of the bond.
    df is a DataFrame of estimated zero prices and index of maturity
    as timestamp and a column 'Zero Prices'.
    bond_type 'Government' actual/actual; other values 30/360
    settlement (datetime): The settlement date for the calculation.

  Returns:
    float: The calculated par yield.
  """


  from datetime import datetime
  import numpy as np
  import pandas as pd

#Check DataFrame of zero prices
  if not isinstance(df, pd.DataFrame):
    raise TypeError("Input 'df' must be a pandas DataFrame.")
  if 'Zero Prices' not in df.columns:
    raise ValueError("The DataFrame must have a 'Zero Prices' column.")

#Assogm settlement date
  if settlement is None:
        settlement = datetime.today()

  #Generate the bond's cash flow dates and payment amounts
  coupon=0.01
  pay_dates, payments = bond_pay_data(maturity, coupon, settlement=settlement)

  # Find the indices of the bond's payment dates within the available zero-coupon price curve.
  # 'zero_prices' is assumed to be a pandas DataFrame with a DatetimeIndex.
  zero_locations = df.index.get_indexer(pay_dates)

# get_indexer returns -1 if in index.
  if -1 in zero_locations:
    return np.nan

  # Retrieve the corresponding zero-coupon prices (discount factors) for each payment date.
  zeros = df['Zero Prices'].iloc[zero_locations]

  # Get the specific zero-coupon price for the bond's final maturity date.
  zero_maturity = zeros.iloc[-1]

  # Par yield is undefined for a zero-coupon bond, so return 'Not a Number'.


  # Sum the discount factors for all payment dates.
  sum_zeros = np.sum(zeros)

  #next_last_coupon_dates imported from python_for_finance
  #Given maturity and settlement dates as dateimes, returns next and last payment dates as
  #datestimes
  next_date, last_date = next_last_coupon_dates(maturity, settlement=settlement)

  #Ratio of days since and between is the adjustment
  if bond_type == 'Government':
      # Actual/Actual convention
      days_since_last = (settlement - last_date).days
      days_between = (next_date - last_date).days
  #other bonds are 30/360
  else:
      # 30/360 convention
      days_since_last = convert_30_360(last_date, settlement)
      days_between = convert_30_360(last_date, next_date)

  #No accrued interest on payment date
  if settlement == next_date:
      accrued_adjust = 0
  else:
      accrued_adjust = days_since_last / days_between

  # Calculate the par yield using the standard formula.
  # The result is multiplied by 2 to annualize it, assuming semi-annual payments.
  par_yield = (1 - zero_maturity) / (sum_zeros - accrued_adjust) * freq

  # Return the final calculated par yield as NumPy array.
  return par_yield
  
def single_newton_raphson(target_value, function=None, data=None, num_attempts=50,
                          guess=0.0001, tolerance=1e-5):
    """
    Finds a root of a function using the Newton-Raphson method.

    Args:
        target_value (float): The target value we want the function to return.
        function (callable): A function that takes a guess and optional data,
                             and returns a tuple (value, derivative).
        data (any, optional): Additional data to pass to the function.
        num_attempts (int): Maximum number of iterations.
        guess (float): Initial guess for the root.
        tolerance (float): The desired precision of the result.

    Returns:
        tuple: (number_of_guesses, final_guess, final_value)
        str: An error message if no solution is found.
    """
    if function is None:
      raise TypeError("function must be specified")
    import numpy as np
    num_guesses = 0

    for num_guesses in range(num_attempts):
        #Ignore second derivative
        value, derivative,*other = function(guess, data)

        # Calculate how far off we are from the target
        error = value - target_value

        # Check for convergence
        if abs(error) <= tolerance:
            return num_guesses, guess, value

        # Avoid division by zero
        if np.isclose(derivative, 0):
            print("No solution found: derivative is zero.")
            return num_guesses+1, guess, value

        # Newton-Raphson update step
        # We use value - target_value (the error) as f(x)
        # The standard formula is x_n+1 = x_n - f(x_n)/f'(x_n)
        guess = guess - error / derivative
    print("No solution found after {} attempts.".format(num_attempts))
    return num_guesses+1, guess, value

def bond_pv(rates=None, data_dict=None, settlement=None):
  '''Calculates a bond's present value, and its first and second derivatives.

  This function prices a bond by discounting its future cash flows using a
  provided set of continuously compounded interest rates. It also computes the
  first and second derivatives of the present value with respect to the rates,
  which are fundamental inputs for calculating risk measures like dollar duration
  and convexity.

  The time to payment is calculated using an Actual/365.25 day-count convention.
  If the provided array of rates is shorter than the number of payments, the
  last rate is used to extrapolate for all subsequent payments.

  Args:
      rates (float or np.ndarray): The continuously compounded discount rate or an
          array of rates. If a single float is provided, it's applied to all
          cash flows.
      data_dict (dict): A dictionary that must contain the bond's payment data.
          It can optionally specify the settlement date.
          - 'pay_data' (tuple): A required tuple of two NumPy arrays:
            (payment_dates, payment_amounts).
          - 'settlement' (datetime.date): An optional valuation date.
      settlement (datetime.date, optional): The valuation date. This argument is
          used if a settlement date is not found in `data_dict`. It defaults
          to the current system date if not provided elsewhere.

  Returns:
      tuple: A tuple of three floats:
      - Present Value: The bond's price,.
      - First Derivative: The derivative of PV with respect to rates.This is related
        to dollar duration.
      - Second Derivative: The second derivative of PV with respect to rates.
        This is related to the bond's dollar convexity.

  Raises:
      TypeError: If 'pay_data' is not a tuple of two NumPy arrays or if
                 'rates' is not numeric.
      ValueError: If the dates and payments arrays within 'pay_data' are not
                  the same size.
  '''
  # Import necessary libraries
  import numpy as np
  from datetime import date

  # if settlement in dictionary otherwise value passed ir no settlement None
  settlement=data_dict.get('settlement',settlement)
  # settlement date is today if it's not provided
  if settlement is None:
    settlement=date.today()
  pay_data=data_dict['pay_data']

# Validate the structure and types within pay_data
  if not isinstance(pay_data, tuple) or len(pay_data) != 2:
      raise TypeError("'pay_data' must be a tuple of (dates_array, payments_array).")
  if not (isinstance(pay_data[0], np.ndarray) and isinstance(pay_data[1], np.ndarray)):
      raise TypeError("Both items in 'pay_data' must be NumPy arrays.")
  if pay_data[0].size != pay_data[1].size:
      raise ValueError("Dates and payments arrays in 'pay_data' must have the same size.")
  # --- Input Validation ---
    # Convert rates to a NumPy array and validate
  if not hasattr(rates, '__iter__'):
      rates = np.array([rates])
  else:
      rates = np.array(rates)

  if not np.issubdtype(rates.dtype, np.number):
        raise TypeError("'rates' must contain only numeric data.")

  # Calculate the time to each payment in years from the settlement date
  # Convert dates to NumPy datetime64, float difference convert to years
  pay_dates=(np.array(pay_data[0],dtype='datetime64[D]')
             -np.datetime64(settlement)).astype(float)/365.25

  # Ensure there is a discount rate for every payment date
  if rates.size<pay_dates.size:
  #If single rate create an array of that single rate
    if rates.size==1:
      rates=np.full(pay_dates.size,rates[0])
   #If more than one but less then size of pay dates
    else:
      rates=np.append(rates,np.full(pay_dates.size-rates.size,rates[-1]))

  #Calculate pv of each payment and sum to get value
  pv_payments=pay_data[1]*np.exp(-rates*pay_dates)
  value=np.sum(pv_payments)

  #Derivatives for each payment
  derivatives_array=-pay_dates*pv_payments

  #First derivative
  first_derivative=np.sum(derivatives_array)

  #Second derivativ
  second_derivative=np.sum(-pay_dates*derivatives_array)
  return value,first_derivative,second_derivative

def calc_ytm(price,maturity,coupon,guess=0.01,settlement=None,freq=2):
  """
  Calculates the Yield to Maturity (YTM) for a bond.

  YTM is the total annualized return an investor can expect if they hold the bond
  until it matures. This function finds the discount rate that equates the
  present value of the bond's future cash flows to its current market price
  using the Newton-Raphson numerical method.

  Args:
    price (float): The current market price of the bond.
    maturity (date): The bond's maturity date.
    coupon (float): The annual coupon rate of the bond (e.g., 0.05 for 5%).
    guess (float, optional): An initial guess for the YTM. Defaults to 0.01 (1%).
                           A good guess can speed up the calculation.
    settlement (date, optional): The date the bond is purchased.
                                 Defaults to the current date.

  Returns:
    float: The calculated Yield to Maturity (YTM) of the bond.
  """
  from datetime import date
  import numpy as np

  # If the initial guess is not a number (NaN), reset it to a default value.
  if np.isnan(guess):
    guess=0.01

  # If no settlement date is provided, use today's date.
  if settlement is None:
    settlement=date.today()

  # Call bond_pay_data() to get the schedule of all future payments
  pay_data=bond_pay_data(maturity,coupon,settlement=settlement)

  # Create a dictionary to pass necessary data to the bond_pv function
  data_dict={'settlement':settlement,'pay_data':pay_data}

  #Newton-Raphson technique passes data to bond_pv to update estimates
  number_iterations,final_result,final_value=single_newton_raphson(price,
                                                                  function=bond_pv,
                                                                  data=data_dict,
                                                                  guess=guess)

  #Failure to converge causes alert, but final shows value at maximum tries
  return final_result

def fred_access(series_id):
  '''
  fred_access reads the St. Louis FED FRED database
  retrieves a single data series
  returns a dataframe with dates index and column names Values


  inputs are fred api key and series id


  '''    
#Need to import pandas,request, and datetime modules
  import pandas as pd
  import requests
  import datetime
#Values are returned as strings and converted to numbers
 # from python_for_finance import financial_strings_numbers as fsn
  error_message='fred_key is not an attribute of the module...execute the statement python_for_finance.finance=fred_key'
  try: fred_key
  except NameError:
    print(error_message)
    return
#Series key identified from page displaying series
#URL includes the series key and the API key
  url='https://api.stlouisfed.org/fred/series/observations?series_id='\
  +series_id+'&api_key='+fred_key+'&file_type=json'
  json_data = requests.get(url)
  data=json_data.json()
#Convert  dates formatted as a string to datetime values
  dates=[datetime.datetime.strptime(x['date'],'%Y-%m-%d') for x in data['observations']]
#Convert values formatted as strings to floating points
  values=[float(x['value']) for x in data['observations']]
#Create a dataframe from the dictionary {Dates':dates,'Values':values']}}
  data_frame=pd.DataFrame({'Dates':dates,'Values':values})
#Set the dates column as the index of the  dataframe
  data_frame.set_index('Dates',inplace=True)
#Return a dataframe
  return data_frame

 
def calc_duration(maturity, coupon,price=None,
                  rates=None,settlement=None, freq=2):
  '''
  Calculates modified duration for a bond.
  Helper functions:
   calc_ytm, bond_pay_data, bond_pv
  Arguments are:
    price: current price of the bond (required for calc_ytm)
    maturity:
      bond's maturity date-datetime or date object (required for calc_ytm, bond_pay_data)
    coupon: bond's annual coupon of par (reqired for calc_ytm and bond_pay_data)
    settlement: bond's settlement date (optional default to current date)
    freq: payment frequency  (optional defaults to 2 for semi-annual (1,2,4 or 12))
  '''
  from datetime import date
  import numpy as np

  # check setlement
  if settlement is None:
    settlement=date.today()
  # calculation of rates
  # if bond sells for 100 ytm calculated from par yield
  if price == 100:
    rates=2*np.log(1+coupon/200)

  # if rates None or zero, calculate as yield to maturity
  if not rates:
        if price is None:
            raise ValueError("Must provide either 'rates' (yield) or 'price'.")
        # FIX: Use the actual 'price' variable, not hardcoded 100
        rates = calc_ytm(price, maturity, coupon, settlement=settlement)

  # Handle iterable rates (e.g. if user passed a list of rates)
  elif hasattr(rates, '__iter__'):
        rates = rates[0]
  # Sanity check: If rate calculation resulted in <= 0, try recalculating
  if isinstance(rates, (int, float)) and rates <= 0:
         # Fallback to Par yield if calculation failed or input was bad
         rates = calc_ytm(100, maturity, coupon, settlement=settlement)

  # data for bond_pv function
  pay_data=bond_pay_data(maturity,coupon,settlement=settlement)
  data_dict={'pay_data':pay_data,'settlement':settlement}

  # calculate present value and first derivative
  value,derivative_first,_=bond_pv(rates=rates,data_dict=data_dict)

  # calculate duration and return value
  duration=-derivative_first/value
  return duration

  
def calc_convexity(maturity, coupon,price=None,
                  rates=None,settlement=None, freq=2):
  '''
  Calculates convexity for a bond.
  Helper functions:
   calc_ytm, bond_pay_data, bond_pv
  Arguments are:
    price: current price of the bond (required for calc_ytm)
    maturity:
      bond's maturity date-datetime or date object (required for calc_ytm, bond_pay_data)
    coupon: bond's annual coupon of par (reqired for calc_ytm and bond_pay_data)
    settlement: bond's settlement date (optional default to current date)
    freq: payment frequency  (optional defaults to 2 for semi-annual (1,2,4 or 12))
  '''
  from datetime import date
  import numpy as np
  # check setlement
  if settlement is None:
    settlement=date.today()
  # calculation of rates
  # if bond sells for 100 ytm calculated from par yield
  if price == 100:
    rates=2*np.log(1+coupon/200)

  # if rates None or zero, calculate as yield to maturity
  if not rates:
        if price is None:
            raise ValueError("Must provide either 'rates' (yield) or 'price'.")
        # FIX: Use the actual 'price' variable, not hardcoded 100
        rates = calc_ytm(price, maturity, coupon, settlement=settlement)

  # Handle iterable rates (e.g. if user passed a list of rates)
  elif hasattr(rates, '__iter__'):
        rates = rates[0]
  # Sanity check: If rate calculation resulted in <= 0, try recalculating
  if isinstance(rates, (int, float)) and rates <= 0:
         # Fallback to Par yield if calculation failed or input was bad
         rates = calc_ytm(100, maturity, coupon, settlement=settlement)

  # data for bond_pv function
  pay_data=bond_pay_data(maturity,coupon,settlement=settlement)
  data_dict={'pay_data':pay_data,'settlement':settlement}

  # calculate present value and second derivative
  value,_,derivative_second=bond_pv(rates=rates,data_dict=data_dict)

  # calculate convexity and return value
  convexity=derivative_second/value
  return convexity

def FEDInvest(price_date):
  """
    Fetches historical security prices from the FedInvest portal.

    Args:
        price_date (datetime.date): The date for which to retrieve prices.
            Note: Current day is typically available after 1:00 PM ET on business days.


    Returns:
        tuple: (pandas.DataFrame, str) if successful. The DataFrame contains
               security details (CUSIP, Price, Yield), and the string is the
               official "Prices For" date stamp from the site.
        tuple: (str, None) if the request fails or no data is found for the date
                (attempt to fetch current day before 1:00 PM ET).

    Example:
        >>> from datetime import date
        >>> df, stamp = FEDInvest(date(2025, 3, 17))
  """

  def make_date(date_value):
    # datetime64 are conerted
    if not isinstance(date_value,(datetime,date)):
      try:
        date_value=pd.Timestamp(date_value).date()
      except Exception as e: # Catch anything else unexpected
        print(f"wrong type for settlement or maturity {e}")

      date_value=pd.Timestamp(settlement).date()
    # convert timestamps and datetimes to date
    else:
      try:
        date_value=date_value.date()
      except:
        pass
    return date_value

  price_date=make_date(price_date)
  # make share date of prices and settlement date are settlement dates
  price_date=adjust_bond_pay_dates(price_date)[0]
  if price_date > date.today():
    return "price_date is in the future", None, None

  settlement_date=price_date+relativedelta(days=1)
  settlement_date=adjust_bond_pay_dates(settlement_date)

  # URL address of Treasury Direct Select A Date
  url = "https://treasurydirect.gov/GA-FI/FedInvest/selectSecurityPriceDate"

  # Standard headers to look like a real browser
  headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36\
     (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Content-Type": "application/x-www-form-urlencoded"
  }
  #  variable names and type identified from inspecting url
  month=str(price_date.month)
  day=str(price_date.day)
  year=str(price_date.year)

  # payload passed in request post
  payload={'priceDate.month':month,
           'priceDate.day':day,
           'priceDate.year':year,
           "submit": "Show Prices"}

  # fires off form and returns prices for date
  try:
        response = requests.post(url, data=payload, headers=headers, timeout=10)
        response.raise_for_status()
  except requests.exceptions.RequestException as e:
        return f"Connection Error: {e}", None

  # reads the html
  # Pandas recommends to wrap the response in StingIO to make file like
  tables=pd.read_html(StringIO(response.text),match='CUSIP')

  # from inspection there is a single table
  df=tables[0]

  df['MATURITY DATE']=pd.to_datetime(df['MATURITY DATE'])

  # drop rows equal to or less than settlement date
  df_filtered=df[df['MATURITY DATE']>pd.to_datetime(settlement_date[0])]

  return df_filtered, price_date,settlement_date[0]

def clean_FEDInvest(df):

    import pandas as pd
    # Filters for Standard Securities
    keep_rows=df['SECURITY TYPE'].str.contains('bill|note|bond',case=False)
    security_df=df[keep_rows].copy()
 
    # Removes Clutter
    drop_columns=['CUSIP','CALL DATE']
    security_df.drop(columns=drop_columns,inplace=True)

    # Creates a Time-Series Index
    security_df.set_index('MATURITY DATE',inplace=True)
    security_df.index=pd.to_datetime(security_df.index)
    security_df.sort_index(inplace=True)

    # Standardizes Financial Terms
    change_column_names={'RATE':'Coupon',
                         'BUY':'Price Ask',
                         'SELL':'Price Bid'}
    security_df.rename(columns=change_column_names,inplace=True)

    # Formats Numeric Data
    numeric_cols = ['Coupon', 'Price Ask', 'Price Bid', 'YIELD']
    for col in numeric_cols:
        if col in security_df.columns:
            security_df[col] = security_df[col].astype(str).str.replace('%', '', regex=False).astype(float)

    return security_df



def has_business_days(start_date, end_date, inclusive='both'):
    """
    Checks if a given date range contains actual business days,
    excluding weekends and US Federal holidays.
    """
    if start_date is None or end_date is None:
        return False
        
    cal = USFederalHolidayCalendar()
    
    # 1. Get standard weekdays in the gap
    b_days = pd.bdate_range(start=start_date, end=end_date, inclusive=inclusive)
    
    # 2. Get holidays in the gap
    holidays = cal.holidays(start=start_date, end=end_date)
    
    # 3. Drop holidays from the weekdays
    actual_b_days = b_days.drop(holidays, errors='ignore')
    
    # Return True if there is at least 1 valid business day missing
    return len(actual_b_days) > 0


class FredReader:
    def __init__(self, api_key=None, cache_dir="fred"):
        # --- COLAB PERSISTENCE LOGIC ---
        if 'google.colab' in sys.modules:
            print("\u2601\uFE0F Colab environment detected. Mounting Google Drive...")
            from google.colab import drive
            drive.mount('/content/drive')
            self.cache_dir = '/content/drive/MyDrive/FRED'
        else:
            self.cache_dir = cache_dir

        # --- API KEY LOGIC ---
        if api_key:
            self.api_key = api_key
        elif os.environ.get("FRED_KEY"):
            self.api_key = os.environ.get("FRED_KEY")
        else:
            print("If you have a FRED key, enter it; otherwise just press Enter:")
            key_input = getpass.getpass(prompt="> ")
            if key_input.strip():
                self.api_key = key_input.strip()
                os.environ["FRED_KEY"] = self.api_key
                print("Key loaded successfully!")
            else:
                self.api_key = None
                print("No key entered. Defaulting to pandas_datareader.")

        if not os.path.exists(self.cache_dir):
            os.makedirs(self.cache_dir)

    # --- REACTIVE NETWORK SHIELD ---
    def _make_api_request(self, url, max_retries=3):
        """Universal helper to handle API calls and 429 Too Many Requests errors."""
        for attempt in range(max_retries):
            response = requests.get(url)
            
            if response.status_code == 429:
                wait_time = 5 * (attempt + 1)
                print(f"\U0001F6A5 Rate limit hit (HTTP 429)! Sleeping for {wait_time}s (Attempt {attempt + 1}/{max_retries})...")
                time.sleep(wait_time)
                continue
                
            response.raise_for_status()
            return response.json()
            
        raise Exception("\u274C Max retries exceeded. The API is strictly rate-limiting you.")

    # --- PUBLIC WRAPPER METHOD ---
    def get_series(self, series_ids, start_date=None, end_date=None, ttl_days=7):
        """Fetches one or more series and returns them in a single merged DataFrame."""
        if isinstance(series_ids, str):
            series_ids = [series_ids]
        elif not hasattr(series_ids, '__iter__'):
            raise TypeError("series_ids must be a string or an iterable of strings.")

        dataframes = []

        for series_id in series_ids:
            print(f"\n--- Processing {series_id} ---")
            df = self._get_single_series(series_id, start_date=start_date, end_date=end_date, ttl_days=ttl_days)
            if df is not None and not df.empty:
                dataframes.append(df)
            else:
                print(f"\u26A0\uFE0F Warning: No data returned for {series_id}.")

        if dataframes:
            if len(dataframes) == 1:
                return dataframes[0]
            print("\n\U0001F9E9 Merging all series into a single DataFrame...")
            combined_df = pd.concat(dataframes, axis=1, join='outer')
            combined_df.sort_index(inplace=True)
            print("\u2705 Merge complete!")
            return combined_df
        else:
            print("\u274C No data could be retrieved.")
            return None

    # --- CORE WORKHORSE METHOD ---
    def _get_single_series(self, series_id, start_date=None, end_date=None, ttl_days=7):
        import datetime as dt
        series_dir = os.path.join(self.cache_dir, series_id)
        os.makedirs(series_dir, exist_ok=True)

        filepath = os.path.abspath(os.path.join(series_dir, f"{series_id}_fred.csv"))
        metadata_file = os.path.abspath(os.path.join(series_dir, "cache_metadata.json"))

        metadata = {}
        metadata_is_stale = True
        metadata_updated_this_run = False

        # --- NESTED METADATA HELPER ---
        def update_metadata():
            nonlocal metadata, metadata_updated_this_run
            if not self.api_key: return
            
            print(f"\U0001F504 Fetching true metadata for {series_id}...")
            meta_url = f"https://api.stlouisfed.org/fred/series?series_id={series_id}&api_key={self.api_key}&file_type=json"
            try:
                meta_data = self._make_api_request(meta_url)
                if 'seriess' in meta_data and len(meta_data['seriess']) > 0:
                    series_info = meta_data['seriess'][0]
                    metadata['series_inception'] = series_info.get('observation_start')
                    metadata['series_last_observed'] = series_info.get('observation_end')
                    metadata['title'] = series_info.get('title')
                    metadata['frequency'] = series_info.get('frequency')
                    metadata['last_updated'] = dt.datetime.now(dt.timezone.utc).isoformat()
                    
                    with open(metadata_file, 'w') as f:
                        json.dump(metadata, f, indent=4)
                    
                    metadata_updated_this_run = True
                    print("\u23F3 Metadata synced and timestamp updated.")
            except Exception as e:
                print(f"\u26A0\uFE0F Could not update metadata: {e}")

        # --- 1. READ LOCAL METADATA ---
        if os.path.exists(metadata_file):
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
                
            if 'last_updated' in metadata:
                last_updated = dt.datetime.fromisoformat(metadata['last_updated'])
                # If the timestamp is from an old cache and lacks a timezone, assign UTC
                if last_updated.tzinfo is None:
                    last_updated = last_updated.replace(tzinfo=dt.timezone.utc)
                
                days_old = (dt.datetime.now(dt.timezone.utc) - last_updated).days
                
                if days_old < ttl_days:
                    metadata_is_stale = False
                    print(f"\U0001F552 Metadata is fresh ({days_old} days old).")
                else:
                    print(f"\u23F3 Metadata is {days_old} days old (TTL: {ttl_days}). It's stale.")

        # --- 2. SMART PING ---
        if not metadata:
            print(f"\U0001F195 First run for {series_id}. Initializing metadata...")
            update_metadata()
        elif self.api_key and end_date and 'series_last_observed' in metadata and metadata_is_stale:
            if pd.to_datetime(end_date) > pd.to_datetime(metadata['series_last_observed']):
                print(f"🔎 Requested date exceeds known end date. Checking for updates...")
                update_metadata()

        # --- 3. CLAMP DATES ---
        if 'series_inception' in metadata and start_date:
            clamped_start = max(pd.to_datetime(start_date), pd.to_datetime(metadata['series_inception']))
            if pd.to_datetime(start_date) < clamped_start:
                start_date = clamped_start.strftime('%Y-%m-%d')
                print(f'\u26A0\uFE0F Adjusted start date to First Available Data: {start_date}')

        if 'series_last_observed' in metadata and end_date:
            clamped_end = min(pd.to_datetime(end_date), pd.to_datetime(metadata['series_last_observed']))
            if pd.to_datetime(end_date) > clamped_end:
                end_date = clamped_end.strftime('%Y-%m-%d')
                print(f'\u26A0\uFE0F FRED has no data past {end_date}. Adjusted request to match.')

        # --- 4. CHECK LOCAL CACHE ---
        cache_valid = False
        if os.path.exists(filepath):
            df = pd.read_csv(filepath, index_col='DATE', parse_dates=True)
            cache_valid = True

            if not df.empty:
                cache_start = df.index.min()
                cache_end = df.index.max()

                if start_date and pd.to_datetime(start_date) < cache_start:
                    if len(pd.bdate_range(start_date, cache_start, inclusive='left')) > 0:
                        cache_valid = False

                if end_date and pd.to_datetime(end_date) > cache_end:
                    if len(pd.bdate_range(cache_end, end_date, inclusive='right')) > 0:
                        cache_valid = False
            else:
                cache_valid = False

            if cache_valid:
                print(f"\u2705 Loaded {series_id} from local cache.")
                if start_date:
                    df = df[df.index >= pd.to_datetime(start_date)]
                if end_date:
                    df = df[df.index <= pd.to_datetime(end_date)]
                return df
            else:
                print(f"\u267B\uFE0F Cache missing or insufficient. Killing cache for {series_id}...")
                os.remove(filepath)

        # --- 5. API FETCH ---
        print(f"\u2601\uFE0F Fetching fresh {series_id} observations...")
        if not metadata_updated_this_run:
            update_metadata()

        try:
            if self.api_key is None:
                print("\u26A0\uFE0F No API key found. Defaulting to pandas_datareader...")
                df = web.DataReader(series_id, 'fred', start=start_date, end=end_date)
            else:
                url = f"https://api.stlouisfed.org/fred/series/observations?series_id={series_id}&api_key={self.api_key}&file_type=json"
                if start_date: url += f"&observation_start={start_date}"
                if end_date: url += f"&observation_end={end_date}"

                safe_url = url.replace(self.api_key, "HIDDEN_KEY")
                print(f"\U0001F4E1 Sending URL to FRED: {safe_url}")

                data = self._make_api_request(url)

                if 'observations' not in data:
                    print(f"\u274C API Error: {data.get('error_message', 'Check API key/ID.')}")
                    return None

                df = pd.DataFrame(data['observations'])
                df['DATE'] = pd.to_datetime(df['date'])
                df[series_id] = pd.to_numeric(df['value'], errors='coerce')
                df = df.set_index('DATE')
                df = df[[series_id]]

            df.dropna(inplace=True)
            df.to_csv(filepath)
            print(f"Success! Data saved to {filepath}.")
            return df

        except Exception as e:
            print(f"\u274C An error occurred: {e}")
            return None

def secure_key_setup(key_name="FRED_KEY"):
    """
    The master UI for setting up API keys. Handles Colab, Local Jupyter, 
    and Ephemeral (Binder) environments automatically.
    """
    try:
        from IPython.display import clear_output, display, HTML
    except ImportError:
        clear_output = lambda: None
        display = lambda x: print(x)
        HTML = lambda x: x

    # 1. Detect Environment
    in_colab = 'google.colab' in sys.modules
    in_binder = 'BINDER_URL' in os.environ or 'BINDER_PORT' in os.environ

    if in_colab:
        from google.colab import userdata
        
        try:
            # STATE 1: Fully Active
            colab_key = userdata.get(key_name)
            if colab_key:
                os.environ[key_name] = colab_key
                clear_output()
                display(HTML(f"""
                <div style="background-color: #d4edda; padding: 15px; border-radius: 8px; border-left: 6px solid #28a745; max-width: 600px;">
                    <h4 style="margin-top: 0; color: #155724; margin-bottom: 5px;">&#10004;&#65039; Key Already Active!</h4>
                    <p style="margin-top: 0; color: #155724; margin-bottom: 0;">We found <b>{key_name}</b> in your Colab Secrets and securely loaded it into the environment. You are ready to fetch data!</p>
                </div>
                """))
                return
                
        except Exception as e:
            # STATE 2: Locked (NotebookAccessError)
            if "Access" in str(type(e).__name__): 
                clear_output()
                display(HTML(f"""
                <div style="background-color: #fff3cd; padding: 15px; border-radius: 8px; border-left: 6px solid #ffc107; max-width: 600px;">
                    <h4 style="margin-top: 0; color: #856404; margin-bottom: 5px;">&#128273; Activation Required</h4>
                    <p style="margin-top: 0; color: #856404;">We found <b>{key_name}</b> in your Colab Secrets, but this notebook doesn't have permission to use it yet.</p>
                    <ul style="color: #856404; margin-bottom: 0;">
                        <li>Click the <b>Key Icon</b> (&#128273;) on the left sidebar.</li>
                        <li>Find <b>{key_name}</b> and check the box next to <b>"Notebook access"</b>.</li>
                        <li>Run this cell again!</li>
                    </ul>
                </div>
                """))
                return
            pass # STATE 3: Doesn't exist. Pass down to the Wizard.

        # --- STATE 3: THE SETUP WIZARD ---
        display(HTML(f"""
        <div style="font-family: sans-serif; max-width: 600px;">
            <div id="wizard-box" style="background-color: #f8f9fa; padding: 20px; border-radius: 8px; border-left: 6px solid #4285f4;">
                <h3 style="margin-top: 0; color: #1a73e8;">&#128274; Secure Key Setup Wizard</h3>
                
                <div id="step1" style="margin-bottom: 15px;">
                    <p><b>Step 1:</b> Click below to copy the required Secret Name:</p>
                    <button onclick="navigator.clipboard.writeText('{key_name}'); this.innerHTML='&#9989; Copied!'; setTimeout(() => this.innerHTML='&#128203; Copy Name: {key_name}', 2000); document.getElementById('step1').style.opacity='0.5'; document.getElementById('step2').style.display='block';" style="padding: 8px 12px; background: white; color: #1a73e8; border: 2px solid #1a73e8; border-radius: 4px; cursor: pointer; font-weight: bold; font-family: monospace;">&#128203; Copy Name: {key_name}</button>
                </div>
                
                <div id="step2" style="display: none; margin-bottom: 15px; border-top: 1px solid #ddd; padding-top: 15px;">
                    <p><b>Step 2:</b> Prepare the Secret in Colab:</p>
                    <ul style="margin-top: 5px;">
                        <li>Click the <b>Key Icon</b> (&#128273;) on the left sidebar.</li>
                        <li>Click <b>"Add new secret"</b> and paste the Name in the left box.</li>
                    </ul>
                    <button onclick="document.getElementById('step2').style.opacity='0.5'; document.getElementById('step2').style.pointerEvents='none'; document.getElementById('step3').style.display='block';" style="padding: 8px 12px; background: #4285f4; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold;">&#9989; The Name is pasted!</button>
                </div>
                
                <div id="step3" style="display: none; margin-bottom: 15px; border-top: 1px solid #ddd; padding-top: 15px;">
                    <p><b>Step 3:</b> Add your API Key!</p>
                    <ul style="margin-top: 5px;">
                        <li>Copy your actual API Key from the LMS.</li>
                        <li>Paste it into the <b>"Value"</b> box in Colab.</li>
                        <li>Toggle <b>"Notebook access"</b> to ON.</li>
                    </ul>
                    <button onclick="document.getElementById('step3').style.opacity='0.5'; document.getElementById('step3').style.pointerEvents='none'; document.getElementById('step4').style.display='block';" style="padding: 8px 12px; background: #34a853; color: white; border: none; border-radius: 4px; cursor: pointer; font-weight: bold;">&#128640; Done! Lock it in.</button>
                </div>
                
                <div id="step4" style="display: none; margin-top: 15px; color: #137333; font-weight: bold; background: #e6f4ea; padding: 15px; border-radius: 4px; line-height: 1.5;">
                    &#127881; <b>Setup complete!</b><br><br>
                    <span style="color: #0d652d; font-size: 0.95em;">&#128161; <b>Pro Tip:</b> This key is securely saved to your Google account! Please re-run this cell to load it.</span>
                </div>
            </div>
        </div>
        """))
        
    else:
        # --- JUPYTER / BINDER LOGIC ---
        target_file = f".{key_name.lower()}"
        
        # UI: Warning if existing file found (Local only)
        if not in_binder and os.path.exists(target_file):
            display(HTML(f"""<div style="background-color: #fff3cd; padding: 15px; border-radius: 8px; border-left: 6px solid #ffc107; font-family: sans-serif; max-width: 600px; margin-bottom: 10px;"><h4 style="margin-top: 0; color: #856404; margin-bottom: 5px;">&#9888;&#65039; WARNING: Key Already Exists</h4><p style="margin-top: 5px; color: #856404; margin-bottom: 0;">A saved key was already found in your vault.</p></div>"""))
            try:
                confirm = input("Do you want to OVERWRITE it? [y/N]: ")
                if confirm.strip().lower() not in ['yes', 'y']:
                    with open(target_file, "r") as f:
                        os.environ[key_name] = f.read().strip()
                    clear_output()
                    display(HTML("""<div style="margin-top: 10px; color: #155724; font-weight: bold; background: #d4edda; padding: 15px; border-radius: 8px; border-left: 6px solid #28a745; max-width: 600px;">&#10005; Setup cancelled. Your existing key was retained <b>and loaded into the environment!</b></div>"""))
                    return
            except KeyboardInterrupt:
                with open(target_file, "r") as f:
                    os.environ[key_name] = f.read().strip()
                clear_output()
                display(HTML("""<div style="margin-top: 10px; color: #155724; font-weight: bold; background: #d4edda; padding: 15px; border-radius: 8px; border-left: 6px solid #28a745; max-width: 600px;">&#10005; Interrupted. Your existing key was retained <b>and loaded into the environment!</b></div>"""))
                return
            clear_output()

        # UI: The Request Box
        if in_binder:
             display(HTML(f"""<div style="background-color: #e2e3e5; padding: 15px; border-radius: 8px; border-left: 6px solid #6c757d; font-family: sans-serif; max-width: 600px; margin-bottom: 10px;"><h3 style="margin-top: 0; color: #383d41; margin-bottom: 5px;">&#9201;&#65039; Ephemeral Session Setup</h3><p style="margin-top: 0; margin-bottom: 0;">You are running in a temporary session. Please paste your <b>{key_name}</b> below.<br><br><b>Note:</b> This key will only persist as long as this browser session remains active.</p></div>"""))
        else:
             display(HTML(f"""<div style="background-color: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 6px solid #4285f4; font-family: sans-serif; max-width: 600px; margin-bottom: 10px;"><h3 style="margin-top: 0; color: #1a73e8; margin-bottom: 5px;">&#128274; Secure Key Setup</h3><p style="margin-top: 0; margin-bottom: 0;">Please paste your <b>{key_name}</b> below. It will be safely vaulted as a hidden file.</p></div>"""))

        # The Input Loop
        print(f"Enter your {key_name} (or type 'quit' to cancel):")
        while True:
            try:
                key_input = getpass.getpass(prompt="> ")
                if key_input.strip().lower() in ['q', 'quit', 'cancel', 'exit']:
                    clear_output(); print("\n\u26A0\uFE0F Setup cancelled by user. No key was saved."); return 
                if key_input.strip():
                    clear_output(); break 
                clear_output(); print("\u26A0\uFE0F Input cannot be empty. Please paste your key (or type 'quit' to cancel):")
            except KeyboardInterrupt:
                clear_output(); print("\n\u26A0\uFE0F Cell execution interrupted. Setup cancelled."); return

        # INJECT INTO ENVIRONMENT IMMEDIATELY
        os.environ[key_name] = key_input.strip()

        # Save Logic (Skip saving to disk if Binder)
        if in_binder:
            display(HTML(f"""<div style="margin-top: 10px; color: #137333; font-weight: bold; background: #e6f4ea; padding: 15px; border-radius: 8px; border-left: 6px solid #34a853; max-width: 600px;">&#127881; <b>Success! Key securely loaded to environment.</b><br><span style="color: #0d652d; font-size: 0.9em;">(Remember: It will be cleared when this session ends.)</span></div>"""))
        else:
            try:
                with open(target_file, "w") as f: f.write(key_input.strip())
                try: os.chmod(target_file, 0o600)
                except: pass
                display(HTML(f"""<div style="margin-top: 10px; color: #137333; font-weight: bold; background: #e6f4ea; padding: 15px; border-radius: 8px; border-left: 6px solid #34a853; max-width: 600px;">&#127881; <b>Success! Your key has been safely vaulted in <code>{target_file}</code></b><br><span style="color: #0d652d; font-size: 0.9em;">&#128161; <b>Pro Tip:</b> Future notebooks will automatically load it!</span></div>"""))
            except Exception as e:
                clear_output(); print(f"\n\u274C Error saving key: {e}")


def load_key_to_env(key_name="FRED_KEY"):
    """
    Silently loads the API key. If it fails, acts as a router to the Setup UI.
    """
    if os.environ.get(key_name): return

    in_colab = 'google.colab' in sys.modules
    if in_colab:
        try:
            from google.colab import userdata
            colab_key = userdata.get(key_name)
            if colab_key:
                os.environ[key_name] = colab_key
                return
        except: pass 

    dot_file = f".{key_name.lower()}"
    if os.path.exists(dot_file):
        with open(dot_file, "r") as f:
            saved_key = f.read().strip()
            if saved_key:
                os.environ[key_name] = saved_key
                return

    # IF ALL FAILS -> Route to Setup UI
    secure_key_setup(key_name)
    
    # The Colab Trap
    if in_colab:
        raise RuntimeError(f"[\u26A0\uFE0F] Setup required! Please complete the wizard above, then re-run this cell.")
